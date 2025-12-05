"""Somex Processor Service - Application Layer"""
import logging
import zipfile
import io
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
from lxml import etree
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from decimal import Decimal
from datetime import datetime

from src.infrastructure.database.somex_repository import SomexRepository
from src.infrastructure.sftp.somex_sftp_client import SomexSftpClient


class ItemsImporter:
    """Helper class to import items from Excel file"""

    def __init__(self, logger: logging.Logger):
        self.logger = logger

    def import_items_from_excel(self, excel_path: str) -> List[Dict[str, Any]]:
        """
        Import items from Excel file

        Expected columns:
        CodigoItem, Referencia, Descripcion, IdPlan, DescPlan,
        IdMayor, DescripcionPlan, RowIdItem, Categoria

        Args:
            excel_path: Path to Excel file

        Returns:
            List of item dictionaries
        """
        items = []

        try:
            wb = openpyxl.load_workbook(excel_path, read_only=True)
            ws = wb.active

            # Read headers from first row
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)

            self.logger.info(f"Excel headers: {headers}")

            # Map expected columns (case insensitive)
            column_map = {}
            expected_columns = [
                'CodigoItem', 'Referencia', 'Descripcion', 'IdPlan', 'DescPlan',
                'IdMayor', 'DescripcionPlan', 'RowIdItem', 'Categoria'
            ]

            for expected in expected_columns:
                for idx, header in enumerate(headers):
                    if header and expected.lower() == str(header).lower().replace(' ', ''):
                        column_map[expected] = idx
                        break

            self.logger.info(f"Column mapping: {column_map}")

            # Read data rows
            for row in ws.iter_rows(min_row=2, values_only=True):
                item = {
                    'codigo_item': str(row[column_map.get('CodigoItem', 0)] or '').strip(),
                    'referencia': str(row[column_map.get('Referencia', 1)] or '').strip(),
                    'descripcion': str(row[column_map.get('Descripcion', 2)] or '').strip(),
                    'id_plan': str(row[column_map.get('IdPlan', 3)] or '').strip(),
                    'desc_plan': str(row[column_map.get('DescPlan', 4)] or '').strip(),
                    'id_mayor': str(row[column_map.get('IdMayor', 5)] or '').strip(),
                    'descripcion_plan': str(row[column_map.get('DescripcionPlan', 6)] or '').strip(),
                    'row_id_item': str(row[column_map.get('RowIdItem', 7)] or '').strip(),
                    'categoria': str(row[column_map.get('Categoria', 8)] or '').strip(),
                }

                # Only add if has codigo_item
                if item['codigo_item']:
                    items.append(item)

            wb.close()

            self.logger.info(f"Imported {len(items)} items from Excel")

        except Exception as e:
            self.logger.error(f"Error importing items from Excel: {e}")
            raise

        return items


class SomexProcessorService:
    """Service for processing Somex ZIP files and generating Excel reports"""

    # UBL Namespaces
    NAMESPACES = {
        'cac': 'urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2',
        'cbc': 'urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2',
        'ext': 'urn:oasis:names:specification:ubl:schema:xsd:CommonExtensionComponents-2',
    }

    def __init__(
        self,
        repository: SomexRepository,
        logger: logging.Logger,
        output_dir: str = "output/somex"
    ):
        self.repository = repository
        self.logger = logger
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def extract_xmls_from_zip(self, zip_path: str) -> List[Tuple[str, bytes]]:
        """
        Extract XML files from a ZIP archive

        Args:
            zip_path: Path to ZIP file

        Returns:
            List of tuples (filename, xml_content)
        """
        xml_files = []

        try:
            self.logger.info(f"Opening ZIP file: {zip_path}")

            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                all_files = zip_ref.namelist()
                self.logger.info(f"Files in ZIP: {all_files}")

                for file_info in zip_ref.filelist:
                    self.logger.debug(f"Checking file: {file_info.filename}")

                    if file_info.filename.lower().endswith('.xml'):
                        self.logger.info(f"Reading XML: {file_info.filename}")
                        xml_content = zip_ref.read(file_info.filename)
                        self.logger.info(
                            f"XML content size: {len(xml_content)} bytes"
                        )
                        xml_files.append((file_info.filename, xml_content))

            self.logger.info(
                f"Extracted {len(xml_files)} XML files from {Path(zip_path).name}"
            )

        except zipfile.BadZipFile as e:
            self.logger.error(f"Invalid ZIP file {zip_path}: {e}")
            raise
        except Exception as e:
            self.logger.error(f"Error extracting XMLs from ZIP {zip_path}: {e}")
            raise

        return xml_files

    def extract_xmls_from_zip_bytes(
        self,
        zip_content: bytes,
        zip_filename: str
    ) -> List[Tuple[str, bytes]]:
        """
        Extract XML files from ZIP bytes

        Args:
            zip_content: ZIP file content as bytes
            zip_filename: Name of the ZIP file

        Returns:
            List of tuples (filename, xml_content)
        """
        xml_files = []

        try:
            with zipfile.ZipFile(io.BytesIO(zip_content), 'r') as zip_ref:
                for file_info in zip_ref.filelist:
                    if file_info.filename.lower().endswith('.xml'):
                        xml_content = zip_ref.read(file_info.filename)
                        xml_files.append((file_info.filename, xml_content))

            self.logger.info(
                f"Extracted {len(xml_files)} XML files from {zip_filename}"
            )

        except Exception as e:
            self.logger.error(
                f"Error extracting XMLs from ZIP {zip_filename}: {e}"
            )
            raise

        return xml_files

    def parse_invoice_xml(self, xml_content: bytes) -> Optional[Dict[str, Any]]:
        """
        Parse invoice XML and extract required data
        Handles both AttachedDocument (with embedded Invoice) and direct Invoice

        Args:
            xml_content: XML content as bytes

        Returns:
            Dictionary with invoice data
        """
        try:
            self.logger.info(f"Parsing XML content ({len(xml_content)} bytes)")
            tree = etree.fromstring(xml_content)

            # Log root tag
            self.logger.info(f"XML root tag: {tree.tag}")

            # Check if this is an AttachedDocument with embedded Invoice
            if 'AttachedDocument' in tree.tag:
                self.logger.info("Detected AttachedDocument, extracting embedded Invoice")
                tree = self._extract_embedded_invoice(tree)

                if tree is None:
                    self.logger.error("Failed to extract embedded Invoice")
                    return None

                self.logger.info(f"Extracted embedded Invoice, new root: {tree.tag}")

            # Extract invoice data using Somex-specific rules
            invoice_data = self._parse_somex_invoice(tree)

            if invoice_data:
                self.logger.info(f"Invoice number: {invoice_data['invoice_number']}")

            return invoice_data

        except Exception as e:
            self.logger.error(f"Error parsing XML: {e}", exc_info=True)
            return None

    def _extract_embedded_invoice(self, attached_doc_tree) -> Optional[any]:
        """
        Extract embedded Invoice XML from AttachedDocument CDATA

        Args:
            attached_doc_tree: Root element of AttachedDocument

        Returns:
            Root element of embedded Invoice or None
        """
        try:
            # Find the CDATA content in Description
            description_elem = attached_doc_tree.find(
                './/cac:Attachment/cac:ExternalReference/cbc:Description',
                self.NAMESPACES
            )

            if description_elem is None or not description_elem.text:
                self.logger.error("No Description element with CDATA found")
                self.logger.error("AttachedDocument structure:")
                self.logger.error(etree.tostring(attached_doc_tree, pretty_print=True).decode()[:1000])
                return None

            # The CDATA contains the actual Invoice XML
            embedded_xml = description_elem.text.strip()

            self.logger.info(f"Embedded XML length: {len(embedded_xml)} characters")
            self.logger.info(f"Embedded XML first 500 chars: {embedded_xml[:500]}")

            # Validate that it's actually XML and not plain text
            # Check if it starts with XML tag
            if not embedded_xml.startswith('<?xml') and not embedded_xml.startswith('<'):
                self.logger.error(
                    f"CDATA content is not XML, it's plain text: {embedded_xml[:200]}"
                )
                return None

            # Check if it contains "Invoice" tag
            if 'Invoice' not in embedded_xml:
                self.logger.error(
                    f"CDATA content doesn't contain Invoice tag. Content: {embedded_xml[:200]}"
                )
                return None

            # Parse the embedded XML
            invoice_tree = etree.fromstring(embedded_xml.encode('utf-8'))
            self.logger.info(f"Successfully parsed embedded Invoice, root: {invoice_tree.tag}")

            return invoice_tree

        except Exception as e:
            self.logger.error(f"Error extracting embedded invoice: {e}", exc_info=True)
            return None

    def _parse_somex_invoice(self, tree) -> Optional[Dict[str, Any]]:
        """
        Parse Invoice XML using Somex-specific rules

        Args:
            tree: Root element of Invoice

        Returns:
            Dictionary with invoice data
        """
        try:
            # Extract invoice number from OrderReference (NOT from main cbc:ID)
            invoice_number = self._get_text(tree, './/cac:OrderReference/cbc:ID')
            if not invoice_number:
                # Fallback to main ID if no OrderReference
                invoice_number = self._get_text(tree, './/cbc:ID')

            # Extract dates
            invoice_date = self._get_text(tree, './/cbc:IssueDate')
            payment_date = self._get_text(tree, './/cbc:PaymentDueDate')
            if not payment_date:
                payment_date = self._get_text(tree, './/cbc:DueDate')

            # Extract buyer information from ReceiverParty (NOT AccountingCustomerParty)
            buyer_party = tree.find(
                './/cac:ReceiverParty',
                self.NAMESPACES
            )

            buyer_nit = ""
            buyer_name = ""
            municipality = ""

            if buyer_party is not None:
                # NIT from ReceiverParty/PartyTaxScheme/CompanyID
                buyer_nit = self._get_text(
                    buyer_party, './/cac:PartyTaxScheme/cbc:CompanyID'
                )
                # Name from ReceiverParty/PartyTaxScheme/RegistrationName
                buyer_name = self._get_text(
                    buyer_party, './/cac:PartyTaxScheme/cbc:RegistrationName'
                )
                # Municipality from ReceiverParty RegistrationAddress
                municipality = self._get_text(
                    buyer_party,
                    './/cac:PartyTaxScheme/cac:RegistrationAddress/cbc:CityName'
                )

                self.logger.info(f"Buyer extracted: NIT={buyer_nit}, Name={buyer_name}, Muni={municipality}")

            # Create base invoice data
            invoice_data = {
                'invoice_number': invoice_number,
                'invoice_date': invoice_date,
                'payment_date': payment_date,
                'buyer_nit': buyer_nit,
                'buyer_name': buyer_name,
                'seller_nit': '800221724',  # Fijo para Somex
                'seller_name': 'SOMEX S.A.S.',  # Fijo para Somex
                'municipality': municipality,
                'items': []
            }

            # Extract line items using Somex-specific parsing
            lines = tree.findall('.//cac:InvoiceLine', self.NAMESPACES)
            self.logger.info(f"Found {len(lines)} invoice lines for invoice {invoice_number}")

            if len(lines) == 0:
                self.logger.warning("No InvoiceLine elements found!")
                self.logger.warning("XML structure (first 1000 chars):")
                self.logger.warning(etree.tostring(tree, pretty_print=True).decode()[:1000])

            for idx, line in enumerate(lines, 1):
                self.logger.info(f"Parsing Somex line item {idx}/{len(lines)}")
                item = self._parse_somex_line_item(line)
                if item:
                    invoice_data['items'].append(item)
                    self.logger.info(
                        f"  ✓ Product: {item['product_name']}, "
                        f"Code: {item['product_code']}, "
                        f"Qty Original: {item['quantity_original']}, "
                        f"Qty Adjusted: {item['quantity_adjusted']}"
                    )
                else:
                    self.logger.warning(f"  ✗ Failed to parse line item {idx}")

            self.logger.info(
                f"Parsed Somex invoice {invoice_number} "
                f"with {len(invoice_data['items'])} items"
            )

            return invoice_data

        except Exception as e:
            self.logger.error(f"Error parsing Somex invoice: {e}", exc_info=True)
            return None

    def _extract_kilos_from_name(self, product_name: str) -> Optional[Decimal]:
        """
        Extract kilos from product name
        Example: "SAL SOMEX CEBA X 40 KILOS" -> 40

        Args:
            product_name: Product name string

        Returns:
            Kilos as Decimal or None if not found
        """
        import re

        try:
            # Look for pattern "X {number} KILO"
            pattern = r'[Xx]\s*(\d+(?:\.\d+)?)\s*[Kk][Ii][Ll]'
            match = re.search(pattern, product_name)

            if match:
                kilos = Decimal(match.group(1))
                self.logger.debug(f"Extracted {kilos} kilos from '{product_name}'")
                return kilos

            return None

        except Exception as e:
            self.logger.error(f"Error extracting kilos from name: {e}")
            return None

    def _parse_somex_line_item(self, line_element) -> Optional[Dict[str, Any]]:
        """
        Parse a single invoice line item using Somex-specific rules

        Args:
            line_element: InvoiceLine XML element

        Returns:
            Dictionary with item data or None
        """
        try:
            # Product name from Note
            product_name = self._get_text(line_element, './/cbc:Note')
            if not product_name:
                # Fallback to Description
                product_name = self._get_text(line_element, './/cac:Item/cbc:Description')

            # Product code from StandardItemIdentification
            product_code = self._get_text(
                line_element,
                './/cac:Item/cac:StandardItemIdentification/cbc:ID'
            )
            if not product_code:
                # Fallback to SellersItemIdentification
                product_code = self._get_text(
                    line_element,
                    './/cac:Item/cac:SellersItemIdentification/cbc:ID'
                )

            # Original quantity from XML
            quantity_str = self._get_text(line_element, './/cbc:InvoicedQuantity')
            quantity_original = Decimal(quantity_str) if quantity_str else Decimal('0')

            # Extract kilos from product name
            kilos = self._extract_kilos_from_name(product_name)

            # Calculate adjusted quantity (quantity * kilos)
            if kilos:
                quantity_adjusted = quantity_original * kilos
                self.logger.debug(
                    f"Adjusted quantity: {quantity_original} * {kilos} = {quantity_adjusted}"
                )
            else:
                quantity_adjusted = quantity_original
                self.logger.warning(
                    f"Could not extract kilos from '{product_name}', "
                    f"using original quantity"
                )

            # Taxable amount (subtotal before tax)
            taxable_amount_str = self._get_text(
                line_element,
                './/cac:TaxTotal/cac:TaxSubtotal/cbc:TaxableAmount'
            )
            taxable_amount = (
                Decimal(taxable_amount_str) if taxable_amount_str else Decimal('0')
            )

            # Calculate unit price: taxable_amount / adjusted_quantity
            if quantity_adjusted > 0:
                unit_price = taxable_amount / quantity_adjusted
            else:
                unit_price = Decimal('0')

            # Tax percentage
            tax_percent_str = self._get_text(
                line_element,
                './/cac:TaxTotal/cac:TaxSubtotal/cac:TaxCategory/cbc:Percent'
            )
            tax_percentage = (
                Decimal(tax_percent_str) if tax_percent_str else Decimal('0')
            )

            return {
                'product_name': product_name or "",
                'product_code': product_code or "SPN-1",  # Default code
                'quantity': quantity_adjusted,  # Adjusted quantity (with kilos)
                'quantity_original': quantity_original,  # Original from XML
                'quantity_adjusted': quantity_adjusted,  # For clarity
                'unit_of_measure': 'KG',  # Always KG for Somex
                'unit_price': unit_price,
                'tax_percentage': tax_percentage,
                'taxable_amount': taxable_amount,
            }

        except Exception as e:
            self.logger.error(f"Error parsing Somex line item: {e}", exc_info=True)
            return None

    def _parse_line_item(self, line_element) -> Optional[Dict[str, Any]]:
        """Parse a single invoice line item"""
        try:
            # Product information
            product_name = (
                self._get_text(line_element, './/cac:Item/cbc:Description') or
                self._get_text(line_element, './/cac:Item/cbc:Name')
            )

            product_code = self._get_text(
                line_element,
                './/cac:Item/cac:SellersItemIdentification/cbc:ID'
            )

            # Quantity
            quantity_str = self._get_text(line_element, './/cbc:InvoicedQuantity')
            quantity = Decimal(quantity_str) if quantity_str else Decimal('0')

            # Unit of measure
            unit_elem = line_element.find(
                './/cbc:InvoicedQuantity',
                self.NAMESPACES
            )
            unit_of_measure = (
                unit_elem.get('unitCode', '') if unit_elem is not None else ''
            )

            # Price
            price_str = self._get_text(
                line_element, './/cac:Price/cbc:PriceAmount'
            )
            unit_price = Decimal(price_str) if price_str else Decimal('0')

            # Tax percentage
            tax_percent_str = self._get_text(
                line_element,
                './/cac:TaxTotal/cac:TaxSubtotal/cbc:Percent'
            )
            tax_percentage = (
                Decimal(tax_percent_str) if tax_percent_str else Decimal('0')
            )

            return {
                'product_name': product_name or "",
                'product_code': product_code or "",
                'quantity': quantity,
                'unit_of_measure': unit_of_measure,
                'unit_price': unit_price,
                'tax_percentage': tax_percentage,
            }

        except Exception as e:
            self.logger.error(f"Error parsing line item: {e}")
            return None

    def _get_text(self, element, xpath: str) -> str:
        """Get text from XPath"""
        if element is None:
            return ""

        result = element.find(xpath, self.NAMESPACES)
        if result is not None and result.text:
            return result.text.strip()
        return ""

    def format_decimal(self, value: Decimal, decimals: int = 5) -> str:
        """Format decimal with specified decimals and comma separator"""
        # Format with 5 decimals
        formatted = f"{value:.{decimals}f}"
        # Replace dot with comma for decimal separator
        return formatted.replace('.', ',')

    def create_excel_template(
        self,
        invoice_data: Dict[str, Any],
        output_filename: Optional[str] = None
    ) -> str:
        """
        Create Excel file from invoice data using the specified template

        Args:
            invoice_data: Dictionary with invoice data
            output_filename: Optional output filename

        Returns:
            Path to created Excel file
        """
        if not output_filename:
            invoice_number = invoice_data.get('invoice_number', 'unknown')
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_filename = f"somex_invoice_{invoice_number}_{timestamp}.xlsx"

        output_path = self.output_dir / output_filename

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Facturas Somex"

        # Define headers according to specification
        headers = [
            "N° Factura",
            "Nombre Producto",
            "Codigo Subyacente",
            "Unidad Medida en Kg,Un,Lt",
            "Peso",
            "Cantidad (5 decimales - separdor coma)",
            "Precio Unitario (5 decimales - separdor coma)",
            "Valor Total (5 decimales - separdor coma)",
            "Fecha Factura Año-Mes-Dia",
            "Fecha Pago Año-Mes-Dia",
            "Nit Comprador (Existente)",
            "Nombre Comprador",
            "Nit Vendedor (Existente)",
            "Nombre Vendedor",
            "Principal V,C",
            "Municipio (Nombre Exacto de la Ciudad)",
            "Iva (N°%)",
            "Descripción",
            "Activa",
            "Factura Activa",
            "Bodega",
            "Incentivo",
            "Cantidad Original (5 decimales - separdor coma)",
            "Moneda (1,2,3)"
        ]

        # Write headers with formatting
        header_fill = PatternFill(
            start_color="366092",
            end_color="366092",
            fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Write data rows
        row_num = 2
        for item in invoice_data.get('items', []):
            ws.cell(row=row_num, column=1).value = invoice_data.get(
                'invoice_number', ''
            )
            ws.cell(row=row_num, column=2).value = item.get('product_name', '')
            ws.cell(row=row_num, column=3).value = item.get('product_code', '')
            ws.cell(row=row_num, column=4).value = item.get('unit_of_measure', '')

            # Peso (weight)
            weight = item.get('weight')
            if weight:
                ws.cell(row=row_num, column=5).value = self.format_decimal(weight)
            else:
                ws.cell(row=row_num, column=5).value = ""

            # Quantity with 5 decimals and comma separator
            ws.cell(row=row_num, column=6).value = self.format_decimal(
                item.get('quantity', Decimal('0'))
            )

            # Unit price with 5 decimals and comma separator
            ws.cell(row=row_num, column=7).value = self.format_decimal(
                item.get('unit_price', Decimal('0'))
            )

            # Valor Total (quantity * unit_price) with 5 decimals and comma separator
            quantity = item.get('quantity', Decimal('0'))
            unit_price = item.get('unit_price', Decimal('0'))
            total_value = quantity * unit_price
            ws.cell(row=row_num, column=8).value = self.format_decimal(total_value)

            ws.cell(row=row_num, column=9).value = invoice_data.get(
                'invoice_date', ''
            )
            ws.cell(row=row_num, column=10).value = invoice_data.get(
                'payment_date', ''
            )
            ws.cell(row=row_num, column=11).value = invoice_data.get(
                'buyer_nit', ''
            )
            ws.cell(row=row_num, column=12).value = invoice_data.get(
                'buyer_name', ''
            )
            ws.cell(row=row_num, column=13).value = invoice_data.get(
                'seller_nit', ''
            )
            ws.cell(row=row_num, column=14).value = invoice_data.get(
                'seller_name', ''
            )
            ws.cell(row=row_num, column=15).value = ""  # Principal V,C
            ws.cell(row=row_num, column=16).value = invoice_data.get(
                'municipality', ''
            )
            ws.cell(row=row_num, column=17).value = str(
                item.get('tax_percentage', '')
            )
            ws.cell(row=row_num, column=18).value = ""  # Descripción
            ws.cell(row=row_num, column=19).value = ""  # Activa
            ws.cell(row=row_num, column=20).value = ""  # Factura Activa
            ws.cell(row=row_num, column=21).value = ""  # Bodega
            ws.cell(row=row_num, column=22).value = ""  # Incentivo

            # Cantidad Original with 5 decimals and comma separator
            ws.cell(row=row_num, column=23).value = self.format_decimal(
                item.get('quantity', Decimal('0'))
            )

            ws.cell(row=row_num, column=24).value = ""  # Moneda

            row_num += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save workbook
        wb.save(output_path)
        self.logger.info(f"Excel file created: {output_path}")

        return str(output_path)

    def process_zip_file(
        self,
        zip_path: str,
        zip_filename: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Process a ZIP file: extract XMLs and parse them

        Args:
            zip_path: Path to ZIP file
            zip_filename: Optional ZIP filename for logging

        Returns:
            Dictionary with processing results including parsed invoices
        """
        if not zip_filename:
            zip_filename = Path(zip_path).name

        results = {
            'zip_filename': zip_filename,
            'total_xmls': 0,
            'processed_xmls': 0,
            'skipped_xmls': 0,
            'failed_xmls': 0,
            'invoices': []  # List of parsed invoice data
        }

        try:
            # Extract XMLs from ZIP
            xml_files = self.extract_xmls_from_zip(zip_path)
            results['total_xmls'] = len(xml_files)

            for xml_filename, xml_content in xml_files:
                # Check if already processed
                if self.repository.is_xml_processed(xml_content):
                    self.logger.info(f"Skipping already processed XML: {xml_filename}")
                    results['skipped_xmls'] += 1
                    continue

                # Parse XML
                invoice_data = self.parse_invoice_xml(xml_content)

                if not invoice_data:
                    self.logger.warning(f"Failed to parse XML: {xml_filename}")
                    results['failed_xmls'] += 1
                    continue

                # Add metadata to invoice data
                invoice_data['xml_filename'] = xml_filename
                invoice_data['xml_content'] = xml_content
                invoice_data['zip_filename'] = zip_filename

                results['invoices'].append(invoice_data)
                results['processed_xmls'] += 1

                self.logger.info(
                    f"Parsed XML {xml_filename}: "
                    f"Invoice {invoice_data.get('invoice_number', 'N/A')} "
                    f"with {len(invoice_data.get('items', []))} items"
                )

        except Exception as e:
            self.logger.error(f"Error processing ZIP file {zip_filename}: {e}")
            raise

        self.logger.info(
            f"=== process_zip_file summary for {zip_filename} ==="
        )
        self.logger.info(f"Total invoices parsed: {len(results['invoices'])}")
        for inv in results['invoices']:
            self.logger.info(
                f"  - {inv.get('invoice_number', 'N/A')}: "
                f"{len(inv.get('items', []))} items"
            )

        return results

    def create_consolidated_excel(
        self,
        all_invoices: List[Dict[str, Any]],
        output_filename: Optional[str] = None
    ) -> str:
        """
        Create a single Excel file with all invoices from multiple ZIPs

        Args:
            all_invoices: List of invoice data dictionaries
            output_filename: Optional output filename

        Returns:
            Path to created Excel file
        """
        # CRITICAL LOGGING: Check what we received
        self.logger.info(f"=== create_consolidated_excel called ===")
        self.logger.info(f"Total invoices received: {len(all_invoices)}")

        for idx, inv in enumerate(all_invoices, 1):
            items_count = len(inv.get('items', []))
            self.logger.info(
                f"  Invoice {idx}: {inv.get('invoice_number', 'N/A')} "
                f"with {items_count} items"
            )

        if not output_filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_filename = f"somex_facturas_consolidadas_{timestamp}.xlsx"

        output_path = self.output_dir / output_filename

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Facturas Somex"

        # Define headers according to specification
        headers = [
            "N° Factura",
            "Nombre Producto",
            "Codigo Subyacente",
            "Unidad Medida en Kg,Un,Lt",
            "Peso",
            "Cantidad (5 decimales - separdor coma)",
            "Precio Unitario (5 decimales - separdor coma)",
            "Valor Total (5 decimales - separdor coma)",
            "Fecha Factura Año-Mes-Dia",
            "Fecha Pago Año-Mes-Dia",
            "Nit Comprador (Existente)",
            "Nombre Comprador",
            "Nit Vendedor (Existente)",
            "Nombre Vendedor",
            "Principal V,C",
            "Municipio (Nombre Exacto de la Ciudad)",
            "Iva (N°%)",
            "Descripción",
            "Activa",
            "Factura Activa",
            "Bodega",
            "Incentivo",
            "Cantidad Original (5 decimales - separdor coma)",
            "Moneda (1,2,3)"
        ]

        # Write headers with formatting
        header_fill = PatternFill(
            start_color="366092",
            end_color="366092",
            fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Write data rows for all invoices
        row_num = 2
        self.logger.info(f"Starting to write invoice rows...")

        for invoice_data in all_invoices:
            items = invoice_data.get('items', [])
            self.logger.info(
                f"Processing invoice {invoice_data.get('invoice_number', 'N/A')}: "
                f"{len(items)} items"
            )

            for item in items:
                # N° Factura
                ws.cell(row=row_num, column=1).value = invoice_data.get(
                    'invoice_number', ''
                )
                # Nombre Producto
                ws.cell(row=row_num, column=2).value = item.get('product_name', '')
                # Codigo Subyacente
                ws.cell(row=row_num, column=3).value = item.get('product_code', 'SPN-1')
                # Unidad Medida en Kg,Un,Lt (siempre KG para Somex)
                ws.cell(row=row_num, column=4).value = 'KG'

                # Peso (weight)
                weight = item.get('weight')
                if weight:
                    ws.cell(row=row_num, column=5).value = self.format_decimal(weight)
                else:
                    ws.cell(row=row_num, column=5).value = ""

                # Cantidad (5 decimales - separador coma) - AJUSTADA con kilos
                ws.cell(row=row_num, column=6).value = self.format_decimal(
                    item.get('quantity_adjusted', item.get('quantity', Decimal('0')))
                )

                # Precio Unitario (5 decimales - separador coma)
                ws.cell(row=row_num, column=7).value = self.format_decimal(
                    item.get('unit_price', Decimal('0'))
                )

                # Valor Total (5 decimales - separador coma)
                quantity = item.get('quantity_adjusted', item.get('quantity', Decimal('0')))
                unit_price = item.get('unit_price', Decimal('0'))
                total_value = quantity * unit_price
                ws.cell(row=row_num, column=8).value = self.format_decimal(total_value)

                # Fecha Factura Año-Mes-Dia
                ws.cell(row=row_num, column=9).value = invoice_data.get(
                    'invoice_date', ''
                )
                # Fecha Pago Año-Mes-Dia
                ws.cell(row=row_num, column=10).value = invoice_data.get(
                    'payment_date', ''
                )
                # Nit Comprador (Existente)
                ws.cell(row=row_num, column=11).value = invoice_data.get(
                    'buyer_nit', ''
                )
                # Nombre Comprador
                ws.cell(row=row_num, column=12).value = invoice_data.get(
                    'buyer_name', ''
                )
                # Nit Vendedor (Existente) - siempre 800221724 para Somex
                ws.cell(row=row_num, column=13).value = '800221724'
                # Nombre Vendedor - siempre SOMEX S.A.S.
                ws.cell(row=row_num, column=14).value = 'SOMEX S.A.S.'
                # Principal V,C - siempre "V"
                ws.cell(row=row_num, column=15).value = "V"
                # Municipio (Nombre Exacto de la Ciudad)
                ws.cell(row=row_num, column=16).value = invoice_data.get(
                    'municipality', ''
                )
                # Iva (N°%)
                ws.cell(row=row_num, column=17).value = str(
                    item.get('tax_percentage', '')
                )
                # Descripción - vacía
                ws.cell(row=row_num, column=18).value = ""
                # Activa - 1
                ws.cell(row=row_num, column=19).value = "1"
                # Factura Activa - 1
                ws.cell(row=row_num, column=20).value = "1"
                # Bodega - vacía
                ws.cell(row=row_num, column=21).value = ""
                # Incentivo - vacía
                ws.cell(row=row_num, column=22).value = ""

                # Cantidad Original (5 decimales - separador coma) - SIN ajustar
                ws.cell(row=row_num, column=23).value = self.format_decimal(
                    item.get('quantity_original', Decimal('0'))
                )

                # Moneda (1,2,3) - siempre 1
                ws.cell(row=row_num, column=24).value = "1"

                row_num += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save workbook
        wb.save(output_path)
        self.logger.info(
            f"Consolidated Excel created: {output_path} with {row_num - 2} rows"
        )

        return str(output_path)
