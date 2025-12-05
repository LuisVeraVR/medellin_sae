"""Pulgarin Inventory Service - Application Layer"""
import logging
import openpyxl
from pathlib import Path
from typing import Dict, List, Optional
from decimal import Decimal


class PulgarinInventoryItem:
    """Represents a product in the Pulgarin inventory"""

    def __init__(self, codigo: str, descripcion: str, peso: Optional[Decimal], unidad_medida: str):
        self.codigo = codigo
        self.descripcion = descripcion
        self.peso = peso
        self.unidad_medida = unidad_medida
        # Normalized description for comparison (lowercase, trimmed)
        self.descripcion_normalizada = descripcion.strip().lower() if descripcion else ""


class PulgarinInventoryService:
    """Service for managing Pulgarin product inventory"""

    def __init__(self, logger: logging.Logger):
        self.logger = logger
        self.inventory: Dict[str, PulgarinInventoryItem] = {}

    def import_from_excel(self, excel_path: str) -> int:
        """
        Import inventory from Excel file

        Expected columns:
        - Codigo: Product code
        - Descripcion: Product description/name
        - PESO: Product weight
        - U/M: Unit of measure

        Args:
            excel_path: Path to Excel file

        Returns:
            Number of items imported
        """
        try:
            wb = openpyxl.load_workbook(excel_path, read_only=True)
            ws = wb.active

            # Read headers from first row
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)

            self.logger.info(f"Excel headers: {headers}")

            # Map expected columns (case insensitive)
            column_map = {}
            expected_columns = ['Codigo', 'Descripcion', 'PESO', 'U/M']

            for expected in expected_columns:
                for idx, header in enumerate(headers):
                    if header and expected.lower() == str(header).lower().replace(' ', '').replace('/', ''):
                        column_map[expected] = idx
                        break

            self.logger.info(f"Column mapping: {column_map}")

            # Clear existing inventory
            self.inventory.clear()

            # Read data rows
            items_imported = 0
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Extract values
                    codigo = str(row[column_map.get('Codigo', 0)] or '').strip()
                    descripcion = str(row[column_map.get('Descripcion', 1)] or '').strip()
                    peso_str = str(row[column_map.get('PESO', 2)] or '').strip()
                    unidad_medida = str(row[column_map.get('U/M', 3)] or '').strip()

                    # Skip empty rows
                    if not descripcion:
                        continue

                    # Parse weight
                    peso = None
                    if peso_str:
                        try:
                            peso = Decimal(peso_str.replace(',', '.'))
                        except:
                            self.logger.warning(f"Row {row_idx}: Could not parse weight '{peso_str}'")

                    # Create inventory item
                    item = PulgarinInventoryItem(
                        codigo=codigo,
                        descripcion=descripcion,
                        peso=peso,
                        unidad_medida=unidad_medida
                    )

                    # Index by normalized description for easy lookup
                    self.inventory[item.descripcion_normalizada] = item
                    items_imported += 1

                except Exception as e:
                    self.logger.error(f"Row {row_idx}: Error processing row: {e}")
                    continue

            wb.close()

            self.logger.info(f"Imported {items_imported} items from Excel")
            return items_imported

        except Exception as e:
            self.logger.error(f"Error importing inventory from Excel: {e}")
            raise

    def find_by_description(self, product_name: str) -> Optional[PulgarinInventoryItem]:
        """
        Find inventory item by product description/name

        Uses normalized (lowercase, trimmed) comparison

        Args:
            product_name: Product name to search for

        Returns:
            PulgarinInventoryItem if found, None otherwise
        """
        normalized_name = product_name.strip().lower() if product_name else ""
        return self.inventory.get(normalized_name)

    def get_weight(self, product_name: str) -> Optional[Decimal]:
        """
        Get weight for a product by name

        Args:
            product_name: Product name

        Returns:
            Weight as Decimal if found, None otherwise
        """
        item = self.find_by_description(product_name)
        return item.peso if item else None

    def get_unit_of_measure(self, product_name: str) -> Optional[str]:
        """
        Get unit of measure for a product by name

        Args:
            product_name: Product name

        Returns:
            Unit of measure if found, None otherwise
        """
        item = self.find_by_description(product_name)
        return item.unidad_medida if item else None

    def get_stats(self) -> Dict[str, int]:
        """
        Get inventory statistics

        Returns:
            Dictionary with statistics
        """
        total_items = len(self.inventory)
        items_with_weight = sum(1 for item in self.inventory.values() if item.peso is not None)
        items_without_weight = total_items - items_with_weight

        return {
            'total_items': total_items,
            'items_with_weight': items_with_weight,
            'items_without_weight': items_without_weight
        }
