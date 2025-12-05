"""Client Tab Widget - Presentation Layer"""
import os
import logging
from pathlib import Path
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
    QTextEdit, QLabel, QCheckBox, QSpinBox, QGroupBox, QFileDialog
)
from PyQt6.QtCore import QTimer, pyqtSignal
from src.domain.entities.client import Client
from src.domain.entities.processing_result import ProcessingResult
from src.application.services.pulgarin_inventory_service import PulgarinInventoryService


class ClientTab(QWidget):
    """Tab widget for individual client"""

    processing_requested = pyqtSignal(str, bool)  # Signal with client_id, allow_reprocess

    def __init__(self, client: Client, logger: logging.Logger, parent=None):
        super().__init__(parent)
        self.client = client
        self.logger = logger
        self.auto_timer = QTimer()
        self.auto_timer.timeout.connect(self._on_auto_process)

        # Initialize inventory service for Pulgarin
        self.inventory_service = None
        if client.id.lower() == "pulgarin":
            self.inventory_service = PulgarinInventoryService(logger)
            # Try to load default inventory file if exists
            default_inventory_path = Path("data/pulgarin_inventario.xlsx")
            if default_inventory_path.exists():
                try:
                    items_count = self.inventory_service.import_from_excel(str(default_inventory_path))
                    logger.info(f"Auto-loaded Pulgarin inventory: {items_count} items")
                except Exception as e:
                    logger.warning(f"Could not auto-load inventory: {e}")

        self._init_ui()

    def _init_ui(self) -> None:
        """Initialize UI components"""
        layout = QVBoxLayout()

        # Client info
        info_label = QLabel(f"<h2>{self.client.name}</h2>")
        layout.addWidget(info_label)

        # Control buttons
        controls_layout = QHBoxLayout()

        self.process_btn = QPushButton("Procesar Ahora")
        self.process_btn.clicked.connect(self._on_process_now)
        controls_layout.addWidget(self.process_btn)

        self.reprocess_checkbox = QCheckBox("Permitir reprocesar correos ya procesados")
        self.reprocess_checkbox.setToolTip(
            "Si está marcado, procesará todos los correos sin importar si ya fueron procesados antes"
        )
        controls_layout.addWidget(self.reprocess_checkbox)

        self.open_output_btn = QPushButton("Abrir Carpeta Output")
        self.open_output_btn.clicked.connect(self._on_open_output)
        controls_layout.addWidget(self.open_output_btn)

        controls_layout.addStretch()
        layout.addLayout(controls_layout)

        # Pulgarin-specific: Inventory management
        if self.client.id.lower() == "pulgarin":
            inventory_group = QGroupBox("Gestión de Inventario (Pulgarin)")
            inventory_layout = QVBoxLayout()

            # Import button and info
            import_layout = QHBoxLayout()

            self.import_inventory_btn = QPushButton("📁 Importar Inventario desde Excel")
            self.import_inventory_btn.clicked.connect(self._on_import_inventory)
            self.import_inventory_btn.setToolTip(
                "Importar catálogo de productos con columnas: Codigo, Descripcion, PESO, U/M"
            )
            import_layout.addWidget(self.import_inventory_btn)

            self.create_example_btn = QPushButton("📝 Crear Excel de Ejemplo")
            self.create_example_btn.clicked.connect(self._on_create_example_inventory)
            self.create_example_btn.setToolTip(
                "Crear un archivo Excel de ejemplo con la estructura correcta"
            )
            import_layout.addWidget(self.create_example_btn)

            import_layout.addStretch()
            inventory_layout.addLayout(import_layout)

            # Inventory stats
            inventory_stats_layout = QHBoxLayout()

            self.inventory_status_label = QLabel("Estado: No cargado")
            inventory_stats_layout.addWidget(self.inventory_status_label)

            self.inventory_items_label = QLabel("Productos: 0")
            inventory_stats_layout.addWidget(self.inventory_items_label)

            self.inventory_with_weight_label = QLabel("Con peso: 0")
            inventory_stats_layout.addWidget(self.inventory_with_weight_label)

            inventory_stats_layout.addStretch()
            inventory_layout.addLayout(inventory_stats_layout)

            # Info text
            info_text = QLabel(
                "ℹ️ El inventario se usa para asignar automáticamente peso (PESO) y unidad de medida (U/M) "
                "a los productos en las facturas, comparando por nombre de producto."
            )
            info_text.setWordWrap(True)
            info_text.setStyleSheet("color: #666; font-size: 10px; padding: 5px;")
            inventory_layout.addWidget(info_text)

            inventory_group.setLayout(inventory_layout)
            layout.addWidget(inventory_group)

            # Update stats if inventory is already loaded
            if self.inventory_service:
                self._update_inventory_stats()

        # Auto-process controls
        auto_group = QGroupBox("Procesamiento Automático")
        auto_layout = QHBoxLayout()

        self.auto_checkbox = QCheckBox("Modo Automático")
        self.auto_checkbox.stateChanged.connect(self._on_auto_mode_changed)
        auto_layout.addWidget(self.auto_checkbox)

        auto_layout.addWidget(QLabel("Intervalo (minutos):"))

        self.interval_spin = QSpinBox()
        self.interval_spin.setMinimum(1)
        self.interval_spin.setMaximum(1440)  # Max 24 hours
        self.interval_spin.setValue(15)
        auto_layout.addWidget(self.interval_spin)

        auto_layout.addStretch()
        auto_group.setLayout(auto_layout)
        layout.addWidget(auto_group)

        # Statistics
        stats_group = QGroupBox("Estadísticas")
        stats_layout = QHBoxLayout()

        self.emails_label = QLabel("Correos procesados: 0")
        stats_layout.addWidget(self.emails_label)

        self.invoices_label = QLabel("Facturas generadas: 0")
        stats_layout.addWidget(self.invoices_label)

        self.errors_label = QLabel("Errores: 0")
        stats_layout.addWidget(self.errors_label)

        stats_layout.addStretch()
        stats_group.setLayout(stats_layout)
        layout.addWidget(stats_group)

        # Log display
        log_group = QGroupBox("Log de Procesamiento")
        log_layout = QVBoxLayout()

        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setMaximumHeight(200)
        log_layout.addWidget(self.log_text)

        clear_log_btn = QPushButton("Limpiar Log")
        clear_log_btn.clicked.connect(self.log_text.clear)
        log_layout.addWidget(clear_log_btn)

        log_group.setLayout(log_layout)
        layout.addWidget(log_group)

        self.setLayout(layout)

    def _on_process_now(self) -> None:
        """Handle process now button click"""
        allow_reprocess = self.reprocess_checkbox.isChecked()
        if allow_reprocess:
            self.log_message("Iniciando procesamiento manual (REPROCESANDO todos los correos)...")
        else:
            self.log_message("Iniciando procesamiento manual...")
        self.process_btn.setEnabled(False)
        self.processing_requested.emit(self.client.id, allow_reprocess)

    def _on_auto_mode_changed(self, state: int) -> None:
        """Handle auto mode checkbox change"""
        if state:
            interval_minutes = self.interval_spin.value()
            interval_ms = interval_minutes * 60 * 1000
            self.auto_timer.start(interval_ms)
            self.log_message(f"Modo automático activado (cada {interval_minutes} minutos)")
        else:
            self.auto_timer.stop()
            self.log_message("Modo automático desactivado")

    def _on_auto_process(self) -> None:
        """Handle automatic processing trigger"""
        allow_reprocess = self.reprocess_checkbox.isChecked()
        self.log_message("Procesamiento automático iniciado...")
        self.processing_requested.emit(self.client.id, allow_reprocess)

    def _on_open_output(self) -> None:
        """Open output folder in file explorer"""
        output_path = Path("output") / self.client.id
        output_path.mkdir(parents=True, exist_ok=True)

        # Open folder in system file explorer
        if os.name == 'nt':  # Windows
            os.startfile(output_path)
        elif os.name == 'posix':  # macOS and Linux
            os.system(f'xdg-open "{output_path}"')

    def log_message(self, message: str) -> None:
        """Add message to log display"""
        self.log_text.append(message)
        self.logger.info(f"[{self.client.id}] {message}")

    def update_stats(self, result: ProcessingResult) -> None:
        """Update statistics display"""
        self.emails_label.setText(f"Correos procesados: {result.emails_processed}")
        self.invoices_label.setText(f"Facturas generadas: {result.invoices_generated}")
        self.errors_label.setText(f"Errores: {result.errors_count}")

        if result.success:
            self.log_message(f"✓ Procesamiento completado: {result.invoices_generated} facturas")
            if result.output_file:
                self.log_message(f"  Archivo: {result.output_file}")
        else:
            self.log_message(f"✗ Procesamiento con errores")
            for error in result.error_messages:
                self.log_message(f"  Error: {error}")

        self.process_btn.setEnabled(True)

    def processing_finished(self) -> None:
        """Re-enable process button after processing"""
        self.process_btn.setEnabled(True)

    # Pulgarin inventory methods
    def _on_import_inventory(self) -> None:
        """Handle import inventory button click"""
        if not self.inventory_service:
            return

        # Open file dialog
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Seleccionar archivo de inventario",
            str(Path("data")),
            "Excel Files (*.xlsx *.xls);;All Files (*)"
        )

        if not file_path:
            return

        try:
            self.log_message(f"Importando inventario desde: {file_path}")
            items_count = self.inventory_service.import_from_excel(file_path)
            self.log_message(f"✓ Inventario importado exitosamente: {items_count} productos")

            # Update stats
            self._update_inventory_stats()

        except Exception as e:
            self.log_message(f"✗ Error al importar inventario: {str(e)}")
            self.logger.error(f"Error importing inventory: {e}", exc_info=True)

    def _on_create_example_inventory(self) -> None:
        """Handle create example inventory button click"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            # Ask where to save
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Guardar archivo de ejemplo",
                str(Path("data/pulgarin_inventario_ejemplo.xlsx")),
                "Excel Files (*.xlsx);;All Files (*)"
            )

            if not file_path:
                return

            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Inventario Pulgarin"

            # Headers
            headers = ["Codigo", "Descripcion", "PESO", "U/M"]

            # Header styling
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            # Write headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Example data
            example_data = [
                ["PROD-001", "SAL REFINADA X 500 GR", "0.5", "KG"],
                ["PROD-002", "SAL REFINADA X 1000 GR", "1.0", "KG"],
                ["PROD-003", "AZUCAR BLANCA X 500 GR", "0.5", "KG"],
                ["PROD-004", "AZUCAR BLANCA X 1000 GR", "1.0", "KG"],
                ["PROD-005", "ACEITE VEGETAL X 1 LITRO", "0.92", "LT"],
            ]

            # Write data
            for row_num, row_data in enumerate(example_data, start=2):
                for col_num, value in enumerate(row_data, start=1):
                    ws.cell(row=row_num, column=col_num).value = value

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Save
            wb.save(file_path)

            self.log_message(f"✓ Archivo de ejemplo creado: {file_path}")
            self.log_message("  Ahora puedes editar este archivo y agregar tus productos")

            # Ask if want to open
            from PyQt6.QtWidgets import QMessageBox
            reply = QMessageBox.question(
                self,
                "Abrir archivo",
                "¿Deseas abrir el archivo de ejemplo ahora?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )

            if reply == QMessageBox.StandardButton.Yes:
                if os.name == 'nt':  # Windows
                    os.startfile(file_path)
                elif os.name == 'posix':  # macOS and Linux
                    os.system(f'xdg-open "{file_path}"')

        except Exception as e:
            self.log_message(f"✗ Error al crear archivo de ejemplo: {str(e)}")
            self.logger.error(f"Error creating example: {e}", exc_info=True)

    def _update_inventory_stats(self) -> None:
        """Update inventory statistics display"""
        if not self.inventory_service:
            return

        stats = self.inventory_service.get_stats()
        total = stats['total_items']
        with_weight = stats['items_with_weight']
        without_weight = stats['items_without_weight']

        if total > 0:
            self.inventory_status_label.setText("Estado: ✓ Cargado")
            self.inventory_status_label.setStyleSheet("color: green; font-weight: bold;")
        else:
            self.inventory_status_label.setText("Estado: No cargado")
            self.inventory_status_label.setStyleSheet("color: orange;")

        self.inventory_items_label.setText(f"Productos: {total}")
        self.inventory_with_weight_label.setText(f"Con peso: {with_weight}")

        if without_weight > 0:
            self.log_message(f"⚠️ {without_weight} productos sin peso en el inventario")

    def get_inventory_service(self):
        """Get the inventory service (for use by parent)"""
        return self.inventory_service
