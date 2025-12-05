"""
Script to create an example Pulgarin inventory Excel file
Run this to create a sample inventory with the correct structure
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

# Create output directory
output_dir = Path("data")
output_dir.mkdir(exist_ok=True)

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Inventario Pulgarin"

# Headers
headers = ["Codigo", "Descripcion", "PESO", "U/M"]

# Header styling
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

# Write headers
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Example data
example_data = [
    ["PROD-001", "SAL REFINADA X 500 GR", "0.5", "KG"],
    ["PROD-002", "SAL REFINADA X 1000 GR", "1.0", "KG"],
    ["PROD-003", "AZUCAR BLANCA X 500 GR", "0.5", "KG"],
    ["PROD-004", "AZUCAR BLANCA X 1000 GR", "1.0", "KG"],
    ["PROD-005", "AZUCAR MORENA X 500 GR", "0.5", "KG"],
    ["PROD-006", "ACEITE VEGETAL X 1 LITRO", "0.92", "LT"],
    ["PROD-007", "ACEITE VEGETAL X 2 LITROS", "1.84", "LT"],
    ["PROD-008", "ARROZ BLANCO X 500 GR", "0.5", "KG"],
    ["PROD-009", "ARROZ BLANCO X 1000 GR", "1.0", "KG"],
    ["PROD-010", "FRIJOL ROJO X 500 GR", "0.5", "KG"],
    ["PROD-011", "PANELA X 500 GR", "0.5", "KG"],
    ["PROD-012", "PANELA X 1000 GR", "1.0", "KG"],
    ["PROD-013", "HARINA DE TRIGO X 500 GR", "0.5", "KG"],
    ["PROD-014", "HARINA DE TRIGO X 1000 GR", "1.0", "KG"],
    ["PROD-015", "LECHE ENTERA X 1 LITRO", "1.03", "LT"],
]

# Write data
for row_num, row_data in enumerate(example_data, start=2):
    for col_num, value in enumerate(row_data, start=1):
        ws.cell(row=row_num, column=col_num).value = value

# Auto-adjust column widths
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 50)
    ws.column_dimensions[column_letter].width = adjusted_width

# Save
output_path = output_dir / "pulgarin_inventario_ejemplo.xlsx"
wb.save(output_path)

print(f"✓ Created example inventory at: {output_path}")
print(f"  - {len(example_data)} example products")
print(f"\nYou can:")
print(f"  1. Open this file in Excel")
print(f"  2. Replace the example data with your actual inventory")
print(f"  3. Save as 'pulgarin_inventario.xlsx'")
print(f"  4. Use it with the PulgarinInventoryService")
